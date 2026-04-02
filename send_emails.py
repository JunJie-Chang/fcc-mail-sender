import os
import re
import smtplib
import time
import logging
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from dotenv import load_dotenv
import openpyxl

load_dotenv()

# ── 設定 ────────────────────────────────────────────────────────
GMAIL_USER     = os.getenv("GMAIL_USER")
GMAIL_APP_PASS = os.getenv("GMAIL_APP_PASSWORD")
SENDER_NAME    = "藍濤亞洲 FCC Partners"
SUBJECT        = "【活動提醒通知】2026/4/8 (三) 13:30【台美新創合作論壇】敬請準時出席"
XLSX_FILE      = "contacts.xlsx"
TEMPLATE_FILE  = "template.html"
MAX_SEND       = 500  # Gmail SMTP 免費帳號每日上限
SLEEP_EVERY    = 10   # 每幾封暫停一次
SLEEP_SECONDS  = 1    # 暫停秒數

# xlsx 欄位索引（1-based）
COL_NUMBER = 1   # A 欄
COL_EMAIL  = 4   # D 欄
COL_NAME   = 5   # E 欄
COL_STATUS = 6   # F 欄（程式自動新增）
# ────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


def load_template() -> str:
    with open(TEMPLATE_FILE, encoding="utf-8") as f:
        return f.read()


def is_valid_email(email: str) -> bool:
    return bool(re.match(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", str(email)))


def build_message(to_name: str, to_email: str, to_number: str, html_body: str) -> MIMEMultipart:
    msg = MIMEMultipart("alternative")
    msg["Subject"] = SUBJECT
    msg["From"]    = f"{SENDER_NAME} <{GMAIL_USER}>"
    msg["To"]      = to_email
    personalised   = html_body.replace("{{name}}", to_name).replace("{{number}}", to_number)
    msg.attach(MIMEText(personalised, "html", "utf-8"))
    return msg


def ensure_status_header(ws) -> None:
    """確保 F1 有 status 標題"""
    if ws.cell(row=1, column=COL_STATUS).value != "status":
        ws.cell(row=1, column=COL_STATUS).value = "status"


def main():
    if not GMAIL_USER or not GMAIL_APP_PASS:
        log.error("請先在 .env 填入 GMAIL_USER 和 GMAIL_APP_PASSWORD")
        return

    html    = load_template()
    wb_data = openpyxl.load_workbook(XLSX_FILE, data_only=True)  # 讀公式計算結果
    wb_write = openpyxl.load_workbook(XLSX_FILE)                 # 寫狀態（保留公式）
    ws_data  = wb_data.active
    ws_write = wb_write.active
    ensure_status_header(ws_write)

    sent    = 0
    failed  = 0
    skipped = 0

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(GMAIL_USER, GMAIL_APP_PASS)
            log.info("SMTP 登入成功")

            for row_idx in range(2, ws_data.max_row + 1):
                if sent >= MAX_SEND:
                    log.info(f"已達本次上限 {MAX_SEND} 封，停止發送。")
                    break

                status = ws_write.cell(row=row_idx, column=COL_STATUS).value
                status = str(status).strip().lower() if status else ""

                if status in ("sent", "error"):
                    skipped += 1
                    continue

                name   = str(ws_data.cell(row=row_idx, column=COL_NAME).value or "").strip()
                email  = str(ws_data.cell(row=row_idx, column=COL_EMAIL).value or "").strip()
                number = str(ws_data.cell(row=row_idx, column=COL_NUMBER).value or "").strip()

                if not is_valid_email(email):
                    ws_write.cell(row=row_idx, column=COL_STATUS).value = "error"
                    failed += 1
                    log.warning(f"[SKIP]  無效 Email，第 {row_idx} 行：{name} <{email}>")
                    wb_write.save(XLSX_FILE)
                    continue

                try:
                    msg = build_message(name, email, number, html)
                    smtp.sendmail(GMAIL_USER, email, msg.as_string())
                    ws_write.cell(row=row_idx, column=COL_STATUS).value = "sent"
                    sent += 1
                    log.info(f"[OK]    {name} <{email}>")

                except Exception as e:
                    ws_write.cell(row=row_idx, column=COL_STATUS).value = "error"
                    failed += 1
                    log.error(f"[FAIL]  {name} <{email}> — {e}")

                finally:
                    wb_write.save(XLSX_FILE)  # 每封發完立即存檔，避免中斷遺失進度

                if sent % SLEEP_EVERY == 0 and sent > 0:
                    time.sleep(SLEEP_SECONDS)

    except smtplib.SMTPAuthenticationError:
        log.error("SMTP 認證失敗，請確認 App Password 是否正確。")
        return

    # ── 摘要 ────────────────────────────────────────────────────
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log.info("=" * 50)
    log.info(f"執行時間：{timestamp}")
    log.info(f"成功：{sent} 封 ／ 失敗：{failed} 封 ／ 略過：{skipped} 封")
    log.info("=" * 50)

    # 在 xlsx 最後一行空一行後寫入摘要
    summary_row = ws_write.max_row + 2
    ws_write.cell(row=summary_row, column=1).value = f"執行時間：{timestamp}"
    ws_write.cell(row=summary_row, column=2).value = f"成功：{sent} 封 ／ 失敗：{failed} 封 ／ 略過：{skipped} 封"
    wb_write.save(XLSX_FILE)


if __name__ == "__main__":
    main()
